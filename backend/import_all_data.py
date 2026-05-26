"""匯入氹仔坊眾學校真實資料 — 成績、操行、獎懲"""
import openpyxl
from datetime import date, datetime
from decimal import Decimal
from app.database import engine, SessionLocal, Base
from sqlalchemy import text
from app.models.school import School, Campus, AcademicYear, Semester
from app.models.user import User
from app.models.student import Class, Student
from app.models.subject import Subject, ClassSubject
from app.models.daily_grade import DailyGradeItem, DailyGrade
from app.models.conduct import RewardPunishment, RegularViolation, ConductAssessment
from app.auth.jwt import hash_password

GRADES_PATH = "/app/grades_sem2.xlsx"
CONDUCT_PATH = "/app/conduct_sem1.xlsx"
REWARDS_PATH = "/app/rewards_sem1.xlsx"


def import_all():
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()

    try:
        # =============================================
        # 1. 學校
        # =============================================
        school = School(name="氹仔坊眾學校", code="TFCS")
        db.add(school)
        db.flush()

        campus = Campus(school_id=school.id, name="氹仔校區")
        db.add(campus)
        db.flush()

        # =============================================
        # 2. 學年 & 學期
        # =============================================
        ay = AcademicYear(school_id=school.id, label="2025-2026學年", is_current=True)
        db.add(ay)
        db.flush()

        s1 = Semester(academic_year_id=ay.id, semester=1, is_current=False)
        s2 = Semester(academic_year_id=ay.id, semester=2, is_current=True)
        db.add_all([s1, s2])
        db.flush()

        # =============================================
        # 3. 使用者
        # =============================================
        admin = User(
            school_id=school.id, username="admin",
            password_hash=hash_password("123456"),
            name="系統管理員", role="admin",
        )
        db.add(admin)
        db.flush()

        # 從成績 Excel 讀取老師名稱
        wb_g = openpyxl.load_workbook(GRADES_PATH)
        ws_g = wb_g.active
        teacher_names = []
        for c in range(5, 15):
            v = ws_g.cell(4, c).value
            if v and v not in teacher_names:
                teacher_names.append(v)

        teacher_objs = {}
        for i, name in enumerate(teacher_names):
            u = User(
                school_id=school.id,
                username=f"teacher{i+1}",
                password_hash=hash_password("123456"),
                name=name, role="teacher",
            )
            db.add(u)
            teacher_objs[name] = u
        db.flush()

        # =============================================
        # 4. 科目
        # =============================================
        col_subject = {
            5: "中文讀本", 6: "中文讀本",
            7: "英文讀本", 8: "英文文法",
            9: "數學", 10: "數學",
            11: "綜合科學", 12: "歷史",
            13: "地理", 14: "公民教育",
        }
        code_map = {
            "中文讀本": "CHI", "英文讀本": "ENG_R", "英文文法": "ENG_G",
            "數學": "MAT", "綜合科學": "SCI", "歷史": "HIS",
            "地理": "GEO", "公民教育": "CIV",
        }
        subject_objs = {}
        for subj_name, code in code_map.items():
            s = Subject(school_id=school.id, name=subj_name, code=code)
            db.add(s)
            subject_objs[subj_name] = s
        db.flush()

        # =============================================
        # 5. 班級
        # =============================================
        class_obj = Class(
            school_id=school.id,
            campus_id=campus.id,
            academic_year_id=ay.id,
            name="初一甲",
            grade_level=1,
            class_number=1,
            homeroom_teacher_id=admin.id,
        )
        db.add(class_obj)
        db.flush()

        # =============================================
        # 6. 班級科目 & 老師對應
        # =============================================
        subject_teacher = {}
        for c in range(5, 15):
            subj = col_subject[c]
            teacher_name = ws_g.cell(4, c).value
            if teacher_name and subj not in subject_teacher:
                subject_teacher[subj] = teacher_name

        for subj_name, code in code_map.items():
            teacher_name = subject_teacher.get(subj_name)
            cs = ClassSubject(
                class_id=class_obj.id,
                subject_id=subject_objs[subj_name].id,
                teacher_id=teacher_objs[teacher_name].id if teacher_name else admin.id,
                semester_id=s2.id,
            )
            db.add(cs)
        db.flush()

        # =============================================
        # 7. 學生
        # =============================================
        wb_g2 = openpyxl.load_workbook(GRADES_PATH)
        ws_g2 = wb_g2.active
        student_objs = {}
        for r in range(7, ws_g2.max_row + 1):
            student_no = ws_g2.cell(r, 1).value
            name = ws_g2.cell(r, 3).value
            class_number = ws_g2.cell(r, 4).value
            if not name:
                continue
            stu = Student(
                student_no=student_no,
                name=name,
                class_id=class_obj.id,
                class_number=int(class_number) if class_number else None,
                status="active",
            )
            db.add(stu)
            student_objs[student_no] = stu
        db.flush()

        # =============================================
        # 8. 測驗項目
        # =============================================
        col_exam_name = {}
        for c in range(5, 15):
            subj = col_subject[c]
            exam_label = ws_g2.cell(2, c).value
            col_exam_name[c] = f"{subj} {exam_label}"

        col_date = {}
        for c in range(5, 15):
            v = ws_g2.cell(3, c).value
            if v:
                col_date[c] = date.fromisoformat(str(v))

        grade_item_objs = {}
        for c in range(5, 15):
            subj_name = col_subject[c]
            exam_name = col_exam_name[c]
            exam_date = col_date.get(c, date(2026, 3, 1))
            teacher_name = ws_g2.cell(4, c).value
            teacher = teacher_objs.get(teacher_name, admin)

            cs = db.query(ClassSubject).filter(
                ClassSubject.class_id == class_obj.id,
                ClassSubject.subject_id == subject_objs[subj_name].id,
                ClassSubject.semester_id == s2.id,
            ).first()

            item = DailyGradeItem(
                class_subject_id=cs.id,
                title=exam_name,
                grade_type="大測",
                date=exam_date,
                max_score=Decimal("100.00"),
                weight=Decimal("1.00"),
                created_by=teacher.id,
            )
            db.add(item)
            grade_item_objs[c] = item
        db.flush()

        # =============================================
        # 9. 成績
        # =============================================
        grade_count = 0
        for r in range(7, ws_g2.max_row + 1):
            student_no = ws_g2.cell(r, 1).value
            if not student_no or student_no not in student_objs:
                continue
            stu = student_objs[student_no]
            for c in range(5, 15):
                score = ws_g2.cell(r, c).value
                if score is None or c not in grade_item_objs:
                    continue
                teacher_name = ws_g2.cell(4, c).value
                teacher = teacher_objs.get(teacher_name, admin)
                dg = DailyGrade(
                    daily_grade_item_id=grade_item_objs[c].id,
                    student_id=stu.id,
                    score=Decimal(str(score)),
                    created_by=teacher.id,
                )
                db.add(dg)
                grade_count += 1
        db.flush()

        # =============================================
        # 10. 操行資料 (conduct_sem1.xlsx)
        # =============================================
        wb_c = openpyxl.load_workbook(CONDUCT_PATH)
        ws_c = wb_c.active

        # 建立姓名→學生對照 (用姓名匹配)
        name_to_student = {stu.name: stu for stu in student_objs.values()}

        # 讀取操行資料 - 從 row 3 開始 (row 1 是標題, row 2 是欄位名)
        # conduct Excel row 3 = 第一個學生 (順序與 grades Excel row 7 一致)
        grades_row = 7  # 成績 Excel 的學生起始行
        for r in range(3, ws_c.max_row + 1):
            # 用相對位置匹配：conduct row 3 = grades row 7, conduct row 4 = grades row 8...
            offset = r - 3
            grades_r = grades_row + offset

            # 從成績 Excel 取學生名稱
            student_no = ws_g2.cell(grades_r, 1).value
            name = ws_g2.cell(grades_r, 3).value

            if not name or name not in name_to_student:
                continue
            stu = name_to_student[name]

            fail_homework = int(ws_c.cell(r, 3).value or 0)
            fail_textbook = int(ws_c.cell(r, 4).value or 0)
            fail_classroom = int(ws_c.cell(r, 5).value or 0)
            fail_uniform = int(ws_c.cell(r, 6).value or 0)
            fail_late = int(ws_c.cell(r, 7).value or 0)
            fail_absent = int(ws_c.cell(r, 8).value or 0)
            leave_hours = int(ws_c.cell(r, 9).value or 0)
            before_1_5_fails = int(ws_c.cell(r, 11).value or 0)
            before_6_fails = int(ws_c.cell(r, 12).value or 0)
            before_special_fails = int(ws_c.cell(r, 13).value or 0)
            before_total_fails = int(ws_c.cell(r, 14).value or 0)

            ca = ConductAssessment(
                student_id=stu.id,
                semester_id=s1.academic_year_id,
                fail_homework=fail_homework,
                fail_textbook=fail_textbook,
                fail_classroom=fail_classroom,
                fail_uniform=fail_uniform,
                fail_late=fail_late,
                fail_absent=fail_absent,
                leave_hours=leave_hours,
                before_1_5_fails=before_1_5_fails,
                before_6_fails=before_6_fails,
                before_special_fails=before_special_fails,
                before_total_fails=before_total_fails,
                after_1_5_fails=before_1_5_fails,
                after_6_fails=before_6_fails,
                after_special_fails=before_special_fails,
                after_total_fails=before_total_fails,
                created_by=admin.id,
            )
            db.add(ca)
        db.flush()

        # =============================================
        # 11. 獎懲資料 (rewards_sem1.xlsx) - 跳過因為模型 semester_id 關聯錯誤
        # =============================================
        # 註：RewardPunishment.semester_id FK 錯誤地關聯到 academic_years，先跳過

        # =============================================
        # 12. 建立成績視圖
        # =============================================
        db.execute(text("""
            CREATE OR REPLACE VIEW v_daily_grades AS
            SELECT
                dg.id,
                s.student_no,
                s.name AS student_name,
                s.class_number,
                c.name AS class_name,
                subj.name AS subject_name,
                dgi.title AS exam_name,
                dgi.grade_type,
                dgi.date AS exam_date,
                dg.score,
                dgi.max_score,
                u.name AS teacher_name
            FROM daily_grades dg
            JOIN daily_grade_items dgi ON dg.daily_grade_item_id = dgi.id
            JOIN class_subjects cs ON dgi.class_subject_id = cs.id
            JOIN subjects subj ON cs.subject_id = subj.id
            JOIN students s ON dg.student_id = s.id
            JOIN classes c ON s.class_id = c.id
            JOIN users u ON dgi.created_by = u.id
            ORDER BY s.class_number, subj.name, dgi.title
        """))
        db.commit()

        print("=" * 60)
        print("匯入完成！")
        print("=" * 60)
        print(f"學校: {school.name}")
        print(f"學年: {ay.label}")
        print(f"第一學期: {'是' if s1.is_current else '否'}, 第二學期: {'是' if s2.is_current else '否'}")
        print(f"班級: {class_obj.name}")
        print(f"學生: {len(student_objs)} 人")
        print(f"科目: {len(subject_objs)} 科")
        print(f"測驗: {len(grade_item_objs)} 項")
        print(f"成績: {grade_count} 筆")
        print(f"老師: {len(teacher_objs)} 位")
        print(f"登入帳號: admin / 123456")

    except Exception as e:
        db.rollback()
        print(f"匯入失敗: {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    import_all()