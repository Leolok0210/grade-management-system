"""
Excel 匯出技能
"""
import os
from app.skills.base import BaseSkill, SkillResult, UserContext
from app.models.daily_grade import DailyGradeItem, DailyGrade
from app.models.student import Student
from app.models.subject import ClassSubject

EXPORT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "exports")
os.makedirs(EXPORT_DIR, exist_ok=True)


class ExportExcel(BaseSkill):
    name = "system.export_excel"
    description = "匯出成績為 Excel 檔案供下載。當使用者要求匯出、下載、輸出成績時使用"
    parameters = {
        "type": "object",
        "properties": {
            "class_subject_id": {"type": "integer", "description": "班級科目ID"},
            "grade_type": {"type": "string", "description": "成績類型篩選（可選）"},
            "export_type": {"type": "string", "enum": ["daily", "daily_summary"], "description": "匯出類型：daily=明細，daily_summary=各學生平均"},
        },
        "required": ["class_subject_id", "export_type"],
    }
    required_role = "teacher"

    async def execute(self, params: dict, context: UserContext, db) -> SkillResult:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        class_subject_id = params["class_subject_id"]
        export_type = params["export_type"]

        cs = db.query(ClassSubject).filter(ClassSubject.id == class_subject_id).first()
        if not cs:
            return SkillResult(success=False, message="找不到班級科目設定")

        from app.models.student import Class
        from app.models.subject import Subject
        cls = db.query(Class).filter(Class.id == cs.class_id).first()
        subj = db.query(Subject).filter(Subject.id == cs.subject_id).first()
        cls_name = cls.name if cls else "?"
        subj_name = subj.name if subj else "?"

        wb = Workbook()
        ws = wb.active
        ws.title = f"{cls_name}{subj_name}"

        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        if export_type == "daily":
            query = db.query(DailyGrade).join(DailyGradeItem).filter(
                DailyGradeItem.class_subject_id == class_subject_id
            )
            if params.get("grade_type"):
                query = query.filter(DailyGradeItem.grade_type == params["grade_type"])

            grades = query.order_by(DailyGradeItem.date.desc()).all()

            if not grades:
                return SkillResult(success=True, message="查無成績記錄，無法匯出")

            ws.append([f"{cls_name} {subj_name} 平時成績明細"])
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
            ws["A1"].font = header_font

            headers = ["學生", "項目", "類型", "分數", "日期"]
            ws.append(headers)
            for cell in ws[2]:
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

            for g in grades:
                student = db.query(Student).filter(Student.id == g.student_id).first()
                ws.append([
                    student.name if student else g.student_id,
                    g.item.title,
                    g.item.grade_type,
                    float(g.score),
                    g.item.date.isoformat(),
                ])
                for cell in ws[ws.max_row]:
                    cell.border = thin_border

        elif export_type == "daily_summary":
            query = db.query(DailyGrade).join(DailyGradeItem).filter(
                DailyGradeItem.class_subject_id == class_subject_id
            )
            grades = query.all()

            if not grades:
                return SkillResult(success=True, message="查無成績記錄，無法匯出")

            student_avgs = {}
            for g in grades:
                if g.student_id not in student_avgs:
                    student_avgs[g.student_id] = []
                student_avgs[g.student_id].append(float(g.score))

            ws.append([f"{cls_name} {subj_name} 成績總表"])
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
            ws["A1"].font = header_font

            headers = ["排名", "學生", "平均分數", "成績筆數"]
            ws.append(headers)
            for cell in ws[2]:
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

            rows = []
            for sid, scores in student_avgs.items():
                student = db.query(Student).filter(Student.id == sid).first()
                avg = round(sum(scores) / len(scores), 2)
                rows.append([student.name if student else sid, avg, len(scores)])

            rows.sort(key=lambda x: x[1], reverse=True)
            for rank, row in enumerate(rows, 1):
                ws.append([rank, *row])
                for cell in ws[ws.max_row]:
                    cell.border = thin_border

        # 儲存
        import uuid
        file_id = str(uuid.uuid4())[:8]
        filename = f"{cls_name}_{subj_name}_{export_type}_{file_id}.xlsx"
        filepath = os.path.join(EXPORT_DIR, filename)
        wb.save(filepath)

        return SkillResult(
            success=True,
            message=f"已匯出 Excel 檔案：{filename}，可至 /api/v1/chat/export/{file_id} 下載",
            data={"filename": filename, "file_id": file_id},
        )

    def preview(self, params: dict, context: UserContext) -> str:
        return f"匯出 {params.get('export_type', '')} 成績 Excel"