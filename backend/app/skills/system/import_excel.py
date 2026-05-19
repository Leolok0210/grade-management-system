"""
Excel 批次匯入技能
"""
import os
from decimal import Decimal
from datetime import date
from app.skills.base import BaseSkill, SkillResult, UserContext
from app.models.daily_grade import DailyGradeItem, DailyGrade
from app.models.semester_grade import SemesterGrade
from app.models.student import Student
from app.models.subject import ClassSubject

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "uploads")


class ImportExcel(BaseSkill):
    name = "system.import_excel"
    description = "從已上傳的 Excel 檔案批次匯入成績。使用者需先上傳檔案，再告知匯入方式"
    parameters = {
        "type": "object",
        "properties": {
            "file_id": {"type": "string", "description": "上傳檔案的ID（由 /chat/upload 返回）"},
            "import_type": {"type": "string", "enum": ["daily", "semester"], "description": "匯入類型：daily=平時成績，semester=學期成績"},
            "class_subject_id": {"type": "integer", "description": "班級科目ID"},
            "semester_id": {"type": "integer", "description": "學期ID"},
            "title": {"type": "string", "description": "成績項目標題（平時成績用，如「第三次小考」）"},
            "grade_type": {"type": "string", "description": "成績類型（平時成績用：作業/小考/大測/課堂參與/口試）"},
        },
        "required": ["file_id", "import_type", "class_subject_id", "semester_id"],
    }
    required_role = "teacher"

    async def execute(self, params: dict, context: UserContext, db) -> SkillResult:
        from openpyxl import load_workbook

        file_id = params["file_id"]
        import_type = params["import_type"]
        class_subject_id = params["class_subject_id"]
        semester_id = params["semester_id"]

        # 找檔案
        file_path = None
        for ext in [".xlsx", ".xls", ".csv"]:
            candidate = os.path.join(UPLOAD_DIR, f"{file_id}{ext}")
            if os.path.exists(candidate):
                file_path = candidate
                break

        if not file_path:
            return SkillResult(success=False, message=f"找不到上傳檔案（file_id={file_id}），請先上傳 Excel 檔案")

        try:
            wb = load_workbook(file_path)
            ws = wb.active
        except Exception as e:
            return SkillResult(success=False, message=f"無法讀取 Excel 檔案: {e}")

        # 取得學生列表
        cs = db.query(ClassSubject).filter(ClassSubject.id == class_subject_id).first()
        if not cs:
            return SkillResult(success=False, message="找不到班級科目設定")

        students = db.query(Student).filter(Student.class_id == cs.class_id).all()
        student_map = {s.student_no: s for s in students}
        # 也支援用班內學號匹配
        student_num_map = {s.class_number: s for s in students if s.class_number}

        imported = []
        errors = []

        if import_type == "daily":
            title = params.get("title", ws.title or "匯入成績")
            grade_type = params.get("grade_type", "其他")

            item = DailyGradeItem(
                class_subject_id=class_subject_id,
                title=title,
                grade_type=grade_type,
                date=date.today(),
                created_by=context.user_id,
            )
            db.add(item)
            db.flush()

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 2:
                    continue
                # 第一欄可能是學號(F250001)或班內學號(1,2,3...)
                student_id_raw = row[0]
                score = row[1]
                if score is None:
                    continue

                student = None
                if isinstance(student_id_raw, str) and student_id_raw.startswith("F"):
                    student = student_map.get(student_id_raw)
                elif student_id_raw is not None:
                    # 可能是班內學號（數字）
                    num = int(student_id_raw) if isinstance(student_id_raw, (int, float)) else None
                    if num:
                        student = student_num_map.get(num)

                if not student:
                    errors.append(f"找不到學生 {student_id_raw}")
                    continue

                grade = DailyGrade(
                    daily_grade_item_id=item.id,
                    student_id=student.id,
                    score=Decimal(str(score)),
                    created_by=context.user_id,
                )
                db.add(grade)
                imported.append([student.name, float(score)])

            db.commit()

        elif import_type == "semester":
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 2:
                    continue
                student_id_raw = row[0]
                student = None
                if isinstance(student_id_raw, str) and student_id_raw.startswith("F"):
                    student = student_map.get(student_id_raw)
                elif student_id_raw is not None:
                    num = int(student_id_raw) if isinstance(student_id_raw, (int, float)) else None
                    if num:
                        student = student_num_map.get(num)

                if not student:
                    errors.append(f"找不到學生 {student_id_raw}")
                    continue

                score_data = {"student_id": student.id}
                if len(row) >= 2 and row[1] is not None:
                    score_data["midterm_score"] = row[1]
                if len(row) >= 3 and row[2] is not None:
                    score_data["final_score"] = row[2]
                if len(row) >= 4 and row[3] is not None:
                    score_data["semester_score"] = row[3]

                existing = db.query(SemesterGrade).filter(
                    SemesterGrade.student_id == student.id,
                    SemesterGrade.class_subject_id == class_subject_id,
                    SemesterGrade.semester_id == semester_id,
                ).first()

                if existing:
                    for key in ["midterm_score", "final_score", "semester_score"]:
                        if key in score_data:
                            setattr(existing, key, Decimal(str(score_data[key])))
                else:
                    sg = SemesterGrade(
                        student_id=student.id,
                        class_subject_id=class_subject_id,
                        semester_id=semester_id,
                        status="draft",
                    )
                    for key in ["midterm_score", "final_score", "semester_score"]:
                        if key in score_data:
                            setattr(sg, key, Decimal(str(score_data[key])))
                    db.add(sg)

                imported.append([student.name, score_data.get("semester_score", "-")])

            db.commit()

        msg = f"已匯入 {len(imported)} 筆成績"
        if errors:
            msg += f"，{len(errors)} 筆有誤"

        return SkillResult(
            success=True,
            message=msg,
            data={"imported": len(imported), "errors": len(errors)},
            data_card={
                "type": "table",
                "title": "匯入結果",
                "payload": {
                    "columns": ["學生", "分數"],
                    "rows": imported,
                },
            },
        )

    def preview(self, params: dict, context: UserContext) -> str:
        return f"匯入 {params.get('import_type', '')} 成績"