"""
學生個人操行評估表 (Individual Student Conduct Assessment)
"""
from openpyxl.styles import Font, Alignment

from app.skills.base import BaseSkill, SkillResult, UserContext
from app.skills.conduct.excel_utils import (
    create_workbook, write_header_row, write_data_row,
    merge_title, save_workbook,
)
from app.models.conduct import RewardPunishment, ConductAssessment
from app.models.student import Student


class StudentIndividualReport(BaseSkill):
    name = "conduct.student_individual_report"
    description = "產生學生個人操行評估表，包含獎懲逐筆明細與抵銷統計。當使用者要求查看學生個人操行、操行評估表、個人操行報告時使用"
    parameters = {
        "type": "object",
        "properties": {
            "student_id": {"type": "integer", "description": "學生ID"},
            "semester_id": {"type": "integer", "description": "學期ID"},
        },
        "required": ["student_id"],
    }
    required_role = "teacher"

    async def execute(self, params: dict, context: UserContext, db) -> SkillResult:
        student_id = params["student_id"]
        semester_id = params.get("semester_id")

        student = db.query(Student).filter(Student.id == student_id).first()
        if not student:
            return SkillResult(success=False, message="找不到學生")

        rp_query = db.query(RewardPunishment).filter(RewardPunishment.student_id == student_id)
        if semester_id:
            rp_query = rp_query.filter(RewardPunishment.semester_id == semester_id)
        rp_records = rp_query.all()

        ca_query = db.query(ConductAssessment).filter(ConductAssessment.student_id == student_id)
        if semester_id:
            ca_query = ca_query.filter(ConductAssessment.semester_id == semester_id)
        ca = ca_query.first()

        reward_rows = []
        punish_rows = []

        for rp in rp_records:
            if rp.reward_type and rp.reward_count:
                reward_rows.append([
                    rp.reward_type, rp.reward_count,
                    rp.reward_reason or "",
                    rp.reward_date.isoformat() if rp.reward_date else "",
                ])
            if rp.punishment_type and rp.punishment_count:
                punish_rows.append([
                    rp.punishment_type, rp.punishment_count,
                    rp.punishment_reason or "",
                    rp.punishment_date.isoformat() if rp.punishment_date else "",
                ])

        reward_cols = ["獎勵類型", "次數", "原因", "日期"]
        punish_cols = ["懲罰類型", "次數", "原因", "日期"]

        reward_table = {
            "type": "table",
            "title": "獎勵記錄",
            "payload": {"columns": reward_cols, "rows": reward_rows},
        }
        punish_table = {
            "type": "table",
            "title": "懲罰記錄",
            "payload": {"columns": punish_cols, "rows": punish_rows},
        }

        stats_cols = ["項目", "優點", "小功", "大功", "缺點", "小過", "大過"]
        before_stats, after_stats = [], []
        if ca:
            before_stats = [
                ca.before_rewards, ca.before_minor_awards, ca.before_major_awards,
                ca.before_minor_infractions, ca.before_major_infractions,
            ]
            after_stats = [
                ca.after_rewards, ca.after_minor_awards, ca.after_major_awards,
                ca.after_minor_infractions, ca.after_major_infractions,
            ]

        stats_rows = [
            ["抵銷前", *before_stats],
            ["抵銷後", *after_stats],
        ]
        stats_table = {
            "type": "table",
            "title": "抵銷統計",
            "payload": {"columns": stats_cols, "rows": stats_rows},
        }

        info_rows = [
            ["學號", student.student_no],
            ["姓名", student.name],
            ["操行評級", ca.current_assessment if ca else "未評"],
            ["評語", ca.comment if ca and ca.comment else ""],
        ]
        info_table = {
            "type": "table",
            "title": "學生資料",
            "payload": {"columns": ["項目", "內容"], "rows": info_rows},
        }

        wb, ws = create_workbook(f"{student.name}_操行評估")

        ws.append(["學生個人操行評估表"])
        merge_title(ws, 1, "學生個人操行評估表", 1, 4)

        row = 2
        ws.cell(row=row, column=1, value="學號").font = Font(bold=True)
        ws.cell(row=row, column=2, value=student.student_no)
        ws.cell(row=row, column=3, value="姓名").font = Font(bold=True)
        ws.cell(row=row, column=4, value=student.name)
        row += 1
        ws.cell(row=row, column=1, value="操行評級").font = Font(bold=True)
        ws.cell(row=row, column=2, value=ca.current_assessment if ca else "未評")
        ws.cell(row=row, column=3, value="義工時數").font = Font(bold=True)
        ws.cell(row=row, column=4, value=ca.volunteer_hours if ca else 0)
        row += 1
        ws.cell(row=row, column=1, value="評語").font = Font(bold=True)
        ws.cell(row=row, column=2, value=ca.comment if ca and ca.comment else "")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

        row += 2
        ws.cell(row=row, column=1, value="獎勵記錄").font = Font(bold=True, size=12)
        row += 1
        ws.append(["獎勵類型", "次數", "原因", "日期"])
        write_header_row(ws, row, ["獎勵類型", "次數", "原因", "日期"])
        row += 1
        for rrow in reward_rows:
            ws.append(rrow)
            write_data_row(ws, row, rrow)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="懲罰記錄").font = Font(bold=True, size=12)
        row += 1
        ws.append(["懲罰類型", "次數", "原因", "日期"])
        write_header_row(ws, row, ["懲罰類型", "次數", "原因", "日期"])
        row += 1
        for prow in punish_rows:
            ws.append(prow)
            write_data_row(ws, row, prow)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="抵銷統計").font = Font(bold=True, size=12)
        row += 1
        ws.append(["項目", "優點", "小功", "大功", "缺點", "小過", "大過"])
        write_header_row(ws, row, ["項目", "優點", "小功", "大功", "缺點", "小過", "大過"])
        row += 1
        ws.append(["抵銷前", *before_stats])
        write_data_row(ws, row, ["抵銷前", *before_stats])
        row += 1
        ws.append(["抵銷後", *after_stats])
        write_data_row(ws, row, ["抵銷後", *after_stats])

        filename, file_id = save_workbook(wb, f"{student.name}_操行評估")

        return SkillResult(
            success=True,
            message=f"已產生 {student.name} 的個人操行評估表",
            data={"filename": filename, "file_id": file_id},
            data_card=info_table,
            data_cards=[reward_table, punish_table, stats_table],
        )

    def preview(self, params: dict, context: UserContext) -> str:
        return f"產生學生個人操行評估表"