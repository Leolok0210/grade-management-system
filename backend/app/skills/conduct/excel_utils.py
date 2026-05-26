"""
Excel utilities for conduct report skills.
"""
import os
import uuid

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

EXPORT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "..", "exports")
try:
    os.makedirs(EXPORT_DIR, exist_ok=True)
except OSError:
    pass  # Vercel read-only filesystem

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


def create_workbook(title: str):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    return wb, ws


def write_header_row(ws, row: int, columns: list, start_col: int = 1):
    for col_idx, col_name in enumerate(columns, start=start_col):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN


def write_data_row(ws, row: int, values: list, start_col: int = 1):
    for col_idx, value in enumerate(values, start=start_col):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN


def merge_title(ws, row: int, text: str, start_col: int, end_col: int):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col, value=text)
    cell.font = Font(bold=True, size=12)
    cell.alignment = CENTER_ALIGN


def save_workbook(wb: Workbook, base_name: str):
    file_id = str(uuid.uuid4())[:8]
    filename = f"{base_name}_{file_id}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)
    return filename, file_id