"""匯入氹仔坊眾學校真實成績資料 — 完整對齊 Excel 欄位"""

import openpyxl
from datetime import date
from decimal import Decimal
from app.database import engine, SessionLocal, Base
from app.models.school import School, Campus, AcademicYear, Semester
from app.models.user import User
from app.models.student import Class, Student
from app.models.subject import Subject, ClassSubject
from app.models.daily_grade import DailyGradeItem, DailyGrade
from app.auth.jwt import hash_password

EXCEL_PATH = "/Users/leo/Downloads/氹仔坊眾學校 - 2025-2026學年初一甲班 第二學期 學生測驗成績報告.xlsx"


def import_real_data():
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        # =============================================
        # 1. 學校
        # =============================================
        school = School(name="氹仔坊眾學校", code="TFCS")
        db.add(school)
        db.flush()

        campus = Campus(school_id=school.id, name="氹仔校區")
        db.add(campus)
        db.flush()

        # =============================================
        # 2. 學年 & 學期
        # =============================================
        ay = AcademicYear(school_id=school.id, label="2025-2026學年", is_current=True)
        db.add(ay)
        db.flush()

        s1 = Semester(academic_year_id=ay.id, semester=1, is_current=False)
        s2 = Semester(academic_year_id=ay.id, semester=2, is_current=True)
        db.add_all([s1, s2])
        db.flush()

        # =============================================
        # 3. 使用者
        # =============================================
        admin = User(
            school_id=school.id, username="admin",
            password_hash=hash_password("123456"),
            name="系統管理員", role="admin",
        )
        db.add(admin)
        db.flush()

        # 從 Excel Row 4 讀取老師名稱
        teacher_names = []
        for c in range(5, 15):
            v = ws.cell(4, c).value
            if v and v not in teacher_names:
                teacher_names.append(v)

        teacher_objs = {}
        for i, name in enumerate(teacher_names):
            u = User(
                school_id=school.id,
                username=f"teacher{i+1}",
                password_hash=hash_password("123456"),
                name=name, role="teacher",
            )
            db.add(u)
            teacher_objs[name] = u
        db.flush()

        # =============================================
        # 4. 科目 — 從 Excel Row 1 讀取
        # =============================================
        # Col 5=中文讀本, 6=中文讀本(merged), 7=英文讀本, 8=英文文法,
        # 9=數學, 10=數學(merged), 11=綜合科學, 12=歷史, 13=地理, 14=公民教育
        col_subject = {
            5: "中文讀本", 6: "中文讀本",
            7: "英文讀本", 8: "英文文法",
            9: "數學", 10: "數學",
            11: "綜合科學", 12: "歷史",
            13: "地理", 14: "公民教育",
        }

        # 科目 → 老師（從 Row 4）
        subject_teacher = {}
        for c in range(5, 15):
            subj = col_subject[c]
            teacher_name = ws.cell(4, c).value
            if teacher_name and subj not in subject_teacher:
                subject_teacher[subj] = teacher_name

        subject_objs = {}
        code_map = {
            "中文讀本": "CHI", "英文讀本": "ENG_R", "英文文法": "ENG_G",
            "數學": "MAT", "綜合科學": "SCI", "歷史": "HIS",
            "地理": "GEO", "公民教育": "CIV",
        }
        for subj_name, code in code_map.items():
            s = Subject(school_id=school.id, name=subj_name, code=code)
            db.add(s)
            subject_objs[subj_name] = s
        db.flush()

        # =============================================
        # 5. 班級
        # =============================================
        class_obj = Class(
            school_id=school.id,
            campus_id=campus.id,
            academic_year_id=ay.id,
            name="初一甲",
            grade_level=1,
            class_number=1,
            homeroom_teacher_id=admin.id,
        )
        db.add(class_obj)
        db.flush()

        # =============================================
        # 6. 班級科目
        # =============================================
        for subj_name, code in code_map.items():
            teacher_name = subject_teacher.get(subj_name)
            cs = ClassSubject(
                class_id=class_obj.id,
                subject_id=subject_objs[subj_name].id,
                teacher_id=teacher_objs[teacher_name].id if teacher_name else admin.id,
                semester_id=s2.id,
            )
            db.add(cs)
        db.flush()

        # =============================================
        # 7. 學生 — Col 1=學生編號, Col 2=班級, Col 3=姓名, Col 4=學號
        # =============================================
        student_objs = {}
        for r in range(7, ws.max_row + 1):
            student_no = ws.cell(r, 1).value   # F250001
            name = ws.cell(r, 3).value          # 卞正成
            class_number = ws.cell(r, 4).value  # 1, 2, 3...

            if not name:
                continue

            stu = Student(
                student_no=student_no,
                name=name,
                class_id=class_obj.id,
                class_number=int(class_number) if class_number else None,
                status="active",
            )
            db.add(stu)
            student_objs[student_no] = stu
        db.flush()

        # =============================================
        # 8. 測驗項目 — 從 Excel Row 1-4 讀取
        # =============================================
        col_exam_name = {}
        for c in range(5, 15):
            subj = col_subject[c]
            exam_label = ws.cell(2, c).value  # 大測1 / 大測2
            col_exam_name[c] = f"{subj} {exam_label}"

        col_date = {}
        for c in range(5, 15):
            v = ws.cell(3, c).value
            if v:
                col_date[c] = date.fromisoformat(str(v))

        grade_item_objs = {}
        for c in range(5, 15):
            subj_name = col_subject[c]
            exam_name = col_exam_name[c]
            exam_date = col_date.get(c, date(2026, 3, 1))
            teacher_name = ws.cell(4, c).value
            teacher = teacher_objs.get(teacher_name, admin)

            cs = db.query(ClassSubject).filter(
                ClassSubject.class_id == class_obj.id,
                ClassSubject.subject_id == subject_objs[subj_name].id,
                ClassSubject.semester_id == s2.id,
            ).first()

            item = DailyGradeItem(
                class_subject_id=cs.id,
                title=exam_name,
                grade_type="大測",
                date=exam_date,
                max_score=Decimal("100.00"),
                weight=Decimal("1.00"),
                created_by=teacher.id,
            )
            db.add(item)
            grade_item_objs[c] = item
        db.flush()

        # =============================================
        # 9. 成績 — Col 5-14 的分數
        # =============================================
        grade_count = 0
        for r in range(7, ws.max_row + 1):
            student_no = ws.cell(r, 1).value
            if not student_no or student_no not in student_objs:
                continue

            stu = student_objs[student_no]

            for c in range(5, 15):
                score = ws.cell(r, c).value
                if score is None or c not in grade_item_objs:
                    continue

                teacher_name = ws.cell(4, c).value
                teacher = teacher_objs.get(teacher_name, admin)

                dg = DailyGrade(
                    daily_grade_item_id=grade_item_objs[c].id,
                    student_id=stu.id,
                    score=Decimal(str(score)),
                    created_by=teacher.id,
                )
                db.add(dg)
                grade_count += 1

        db.flush()
        db.commit()

        # =============================================
        # 10. 建立成績視圖（方便查詢）
        # =============================================
        db.execute(text("""
            CREATE OR REPLACE VIEW v_daily_grades AS
            SELECT
                dg.id,
                s.student_no,
                s.name AS student_name,
                s.class_number,
                c.name AS class_name,
                subj.name AS subject_name,
                dgi.title AS exam_name,
                dgi.grade_type,
                dgi.date AS exam_date,
                dg.score,
                dgi.max_score,
                u.name AS teacher_name
            FROM daily_grades dg
            JOIN daily_grade_items dgi ON dg.daily_grade_item_id = dgi.id
            JOIN class_subjects cs ON dgi.class_subject_id = cs.id
            JOIN subjects subj ON cs.subject_id = subj.id
            JOIN students s ON dg.student_id = s.id
            JOIN classes c ON s.class_id = c.id
            JOIN users u ON dgi.created_by = u.id
            ORDER BY s.class_number, subj.name, dgi.title
        """))
        db.commit()

        # =============================================
        # 驗證：對比 Excel vs DB
        # =============================================
        print("=" * 60)
        print("匯入完成！資料驗證：")
        print("=" * 60)
        print(f"學校: {school.name}")
        print(f"學年: {ay.label} 第2學期")
        print(f"班級: {class_obj.name}")
        print(f"學生: {len(student_objs)} 人")
        print(f"科目: {len(subject_objs)} 科")
        print(f"測驗: {len(grade_item_objs)} 項")
        print(f"成績: {grade_count} 筆")
        print(f"老師: {len(teacher_objs)} 位")
        print()

        # 對比前5個學生
        print("Excel vs DB 對比（前5位學生）：")
        for r in range(7, 12):
            student_no = ws.cell(r, 1).value
            excel_name = ws.cell(r, 3).value
            excel_num = ws.cell(r, 4).value
            excel_scores = [ws.cell(r, c).value for c in range(5, 15)]

            if student_no in student_objs:
                stu = student_objs[student_no]
                db_scores = []
                for c in range(5, 15):
                    if c in grade_item_objs:
                        dg = db.query(DailyGrade).filter(
                            DailyGrade.daily_grade_item_id == grade_item_objs[c].id,
                            DailyGrade.student_id == stu.id,
                        ).first()
                        db_scores.append(float(dg.score) if dg else None)

                match = "✓" if excel_scores == db_scores else "✗"
                print(f"  {match} {excel_name} (學號{excel_num}): Excel={excel_scores} DB={db_scores}")

    except Exception as e:
        db.rollback()
        print(f"匯入失敗: {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    import_real_data()