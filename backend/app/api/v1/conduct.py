"""
德育管理 API - 獎懲登記、常規違紀、操行評估
"""
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
from datetime import date
from app.database import get_db
from app.deps import get_current_user
from app.models.user import User
from app.models.conduct import RewardPunishment, RegularViolation, ConductAssessment
from app.models.student import Student, Class

router = APIRouter(prefix="/conduct", tags=["德育管理"])


# ============ 獎懲登記 ============

class RewardPunishmentCreate(BaseModel):
    student_id: int
    semester_id: int
    reward_type: Optional[str] = None
    reward_count: int = 0
    reward_reason: Optional[str] = None
    reward_date: Optional[date] = None
    punishment_type: Optional[str] = None
    punishment_count: int = 0
    punishment_reason: Optional[str] = None
    punishment_date: Optional[date] = None


@router.post("/register")
async def register_reward_punishment(
    data: RewardPunishmentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """登記獎勵或懲罰"""
    student = db.query(Student).filter(Student.id == data.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="學生不存在")

    rp = RewardPunishment(
        student_id=data.student_id,
        semester_id=data.semester_id,
        reward_type=data.reward_type,
        reward_count=data.reward_count,
        reward_reason=data.reward_reason,
        reward_date=data.reward_date,
        punishment_type=data.punishment_type,
        punishment_count=data.punishment_count,
        punishment_reason=data.punishment_reason,
        punishment_date=data.punishment_date,
        created_by=current_user.id,
    )
    db.add(rp)
    db.commit()
    db.refresh(rp)

    return {"message": "登記成功", "id": rp.id}


# ============ 常規違紀登記 ============

class RegularViolationCreate(BaseModel):
    student_id: int
    class_id: int
    semester_id: int
    violation_type: str  # 欠作業/欠課本/上課違規/儀表不符/遲到/缺席/請假
    count: int = 0
    record_date: Optional[date] = None


@router.post("/violations")
async def register_violation(
    data: RegularViolationCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """登記常規違紀"""
    student = db.query(Student).filter(Student.id == data.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="學生不存在")

    rv = RegularViolation(
        student_id=data.student_id,
        class_id=data.class_id,
        semester_id=data.semester_id,
        violation_type=data.violation_type,
        count=data.count,
        record_date=data.record_date,
        created_by=current_user.id,
    )
    db.add(rv)
    db.commit()
    db.refresh(rv)

    return {"message": "登記成功", "id": rv.id}


# ============ 操行評估 ============

class ConductAssessmentCreate(BaseModel):
    student_id: int
    semester_id: int
    fail_homework: int = 0
    fail_textbook: int = 0
    fail_classroom: int = 0
    fail_uniform: int = 0
    fail_late: int = 0
    fail_absent: int = 0
    leave_hours: int = 0
    before_rewards: int = 0
    before_minor_awards: int = 0
    before_major_awards: int = 0
    after_rewards: int = 0
    after_minor_awards: int = 0
    after_major_awards: int = 0
    volunteer_hours: float = 0
    max_assessment: Optional[str] = None
    previous_assessment: Optional[str] = None
    current_assessment: Optional[str] = None
    change: Optional[str] = None
    cumulative_fails: int = 0
    comment: Optional[str] = None


@router.post("/assessment")
async def save_conduct_assessment(
    data: ConductAssessmentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """儲存操行評估"""
    student = db.query(Student).filter(Student.id == data.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="學生不存在")

    # 檢查是否已有記錄
    existing = db.query(ConductAssessment).filter(
        ConductAssessment.student_id == data.student_id,
        ConductAssessment.semester_id == data.semester_id,
    ).first()

    if existing:
        # 更新現有記錄
        for key, value in data.model_dump().items():
            setattr(existing, key, value)
        db.commit()
        return {"message": "更新成功", "id": existing.id}

    # 新增記錄
    ca = ConductAssessment(**data.model_dump(), created_by=current_user.id)
    db.add(ca)
    db.commit()
    db.refresh(ca)

    return {"message": "儲存成功", "id": ca.id}


# ============ 查詢 ============

@router.get("/class/{class_id}")
async def get_class_conduct(
    class_id: int,
    semester_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """取得班級的德育資料"""
    students = db.query(Student).filter(Student.class_id == class_id).order_by(Student.class_number).all()
    student_ids = [s.id for s in students]

    # 取得操行評估
    query = db.query(ConductAssessment).filter(ConductAssessment.student_id.in_(student_ids))
    if semester_id:
        query = query.filter(ConductAssessment.semester_id == semester_id)
    assessments = {a.student_id: a for a in query.all()}

    # 組裝資料
    result = []
    for student in students:
        ca = assessments.get(student.id)
        if ca:
            result.append({
                "student_no": student.student_no,
                "name": student.name,
                "class_number": student.class_number,
                "fail_homework": ca.fail_homework,
                "fail_textbook": ca.fail_textbook,
                "fail_classroom": ca.fail_classroom,
                "fail_uniform": ca.fail_uniform,
                "fail_late": ca.fail_late,
                "fail_absent": ca.fail_absent,
                "leave_hours": ca.leave_hours,
                "before_rewards": ca.before_rewards,
                "before_minor_awards": ca.before_minor_awards,
                "before_major_awards": ca.before_major_awards,
                "after_rewards": ca.after_rewards,
                "after_minor_awards": ca.after_minor_awards,
                "after_major_awards": ca.after_major_awards,
                "volunteer_hours": ca.volunteer_hours,
                "max_assessment": ca.max_assessment,
                "current_assessment": ca.current_assessment,
                "comment": ca.comment,
            })

    return result


@router.get("/rewards/{student_id}")
async def get_student_rewards(
    student_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """取得學生的獎懲記錄"""
    records = db.query(RewardPunishment).filter(
        RewardPunishment.student_id == student_id,
        RewardPunishment.is_active == True,
    ).all()

    return [
        {
            "id": r.id,
            "reward_type": r.reward_type,
            "reward_count": r.reward_count,
            "reward_reason": r.reward_reason,
            "reward_date": r.reward_date.isoformat() if r.reward_date else None,
            "punishment_type": r.punishment_type,
            "punishment_count": r.punishment_count,
            "punishment_reason": r.punishment_reason,
            "punishment_date": r.punishment_date.isoformat() if r.punishment_date else None,
        }
        for r in records
    ]


class ExcelImportRequest(BaseModel):
    file_path: str
    class_id: int
    semester_id: int


@router.post("/excel-import")
async def import_conduct_from_excel(
    req: ExcelImportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """從 Excel 匯入操行資料"""
    import openpyxl

    file_path = req.file_path
    class_id = req.class_id
    semester_id = req.semester_id

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        imported = 0
        for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
            if not row[0]:
                continue

            try:
                class_number = int(row[0]) if row[0] else None
                if class_number is None:
                    continue
            except (ValueError, TypeError):
                continue

            name = row[1] if len(row) > 1 else ""
            fail_homework = int(row[2]) if row[2] else 0
            fail_textbook = int(row[3]) if row[3] else 0
            fail_classroom = int(row[4]) if row[4] else 0
            fail_uniform = int(row[5]) if row[5] else 0
            fail_late = int(row[6]) if row[6] else 0
            fail_absent = int(row[7]) if row[7] else 0
            leave_hours = int(row[8]) if row[8] else 0
            before_rewards = int(row[16]) if row[16] else 0
            before_minor_awards = int(row[17]) if row[17] else 0
            before_major_awards = int(row[18]) if row[18] else 0
            after_rewards = int(row[26]) if row[26] else 0
            after_minor_awards = int(row[27]) if row[27] else 0
            after_major_awards = int(row[28]) if row[28] else 0
            volunteer_hours = float(row[29]) if row[29] else 0
            offset_count = str(row[30]) if row[30] else None
            max_assessment = str(row[31]) if row[31] else None
            previous_assessment = str(row[32]) if row[32] else None
            current_assessment = str(row[33]) if row[33] else None
            change = str(row[34]) if row[34] else None
            cumulative_fails_val = row[35]
            cumulative_fails = int(cumulative_fails_val) if cumulative_fails_val and str(cumulative_fails_val).isdigit() else 0
            comment = str(row[36]) if row[36] else None

            # 找學生
            student = db.query(Student).filter(
                Student.class_id == class_id,
                Student.class_number == class_number,
            ).first()

            if not student:
                continue

            # 檢查是否已有記錄
            existing = db.query(ConductAssessment).filter(
                ConductAssessment.student_id == student.id,
                ConductAssessment.semester_id == semester_id,
            ).first()

            ca_data = {
                "student_id": student.id,
                "semester_id": semester_id,
                "fail_homework": fail_homework,
                "fail_textbook": fail_textbook,
                "fail_classroom": fail_classroom,
                "fail_uniform": fail_uniform,
                "fail_late": fail_late,
                "fail_absent": fail_absent,
                "leave_hours": leave_hours,
                "before_rewards": before_rewards,
                "before_minor_awards": before_minor_awards,
                "before_major_awards": before_major_awards,
                "after_rewards": after_rewards,
                "after_minor_awards": after_minor_awards,
                "after_major_awards": after_major_awards,
                "volunteer_hours": volunteer_hours,
                "offset_count": offset_count,
                "max_assessment": max_assessment,
                "previous_assessment": previous_assessment,
                "current_assessment": current_assessment,
                "change": change,
                "cumulative_fails": cumulative_fails,
                "comment": comment,
            }

            if existing:
                for key, value in ca_data.items():
                    setattr(existing, key, value)
            else:
                existing = ConductAssessment(**ca_data, created_by=current_user.id)
                db.add(existing)

            imported += 1

        db.commit()
        return {"message": f"成功匯入 {imported} 筆記錄"}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class RewardExcelImportRequest(BaseModel):
    file_path: str
    class_id: int
    semester_id: int


@router.post("/rewards-excel-import")
async def import_rewards_punishments_from_excel(
    req: RewardExcelImportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """從 Excel 匯入獎懲資料（獎懲記錄列表格式）"""
    import openpyxl
    import re

    file_path = req.file_path
    class_id = req.class_id
    semester_id = req.semester_id

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        imported = 0
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            # Row format: (學生編號, 班別, 學號, 姓名, 優點項目, 缺點項目)
            if not row[2]:  # 學號
                continue

            class_number = int(row[2]) if row[2] else None
            name = row[3] if len(row) > 3 else ""
            reward_text = row[4] if len(row) > 4 else None
            punishment_text = row[5] if len(row) > 5 else None

            # 解析優點次數
            reward_count = 0
            reward_reason = None
            if reward_text and reward_text != "----":
                # 計算 "按章應記優點X次" 的次數
                matches = re.findall(r'按章應記優點(\d+)次', reward_text)
                reward_count = sum(int(m) for m in matches)
                reward_reason = reward_text

            # 解析缺點次數
            punishment_count = 0
            punishment_reason = None
            if punishment_text and punishment_text != "----":
                # 計算 "按章應記缺點X個" 的個數
                matches = re.findall(r'按章應記缺點(\d+)個', punishment_text)
                punishment_count = sum(int(m) for m in matches)
                punishment_reason = punishment_text

            # 找學生
            student = db.query(Student).filter(
                Student.class_id == class_id,
                Student.class_number == class_number,
            ).first()

            if not student:
                continue

            # 檢查是否已有獎懲記錄
            existing = db.query(RewardPunishment).filter(
                RewardPunishment.student_id == student.id,
                RewardPunishment.semester_id == semester_id,
            ).first()

            if existing:
                # 更新現有記錄
                if reward_count > 0:
                    existing.reward_count = reward_count
                    existing.reward_reason = reward_reason
                    existing.reward_type = "優點"
                if punishment_count > 0:
                    existing.punishment_count = punishment_count
                    existing.punishment_reason = punishment_reason
                    existing.punishment_type = "缺點"
            else:
                # 新增記錄
                rp = RewardPunishment(
                    student_id=student.id,
                    semester_id=semester_id,
                    reward_type="優點" if reward_count > 0 else None,
                    reward_count=reward_count,
                    reward_reason=reward_reason,
                    punishment_type="缺點" if punishment_count > 0 else None,
                    punishment_count=punishment_count,
                    punishment_reason=punishment_reason,
                    created_by=current_user.id,
                )
                db.add(rp)

            imported += 1

        db.commit()
        return {"message": f"成功匯入 {imported} 筆記錄"}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))